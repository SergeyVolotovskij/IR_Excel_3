# import the main libraries
from openpyxl import load_workbook
import pyexcel
from openpyxl.styles import Font, colors
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
from openpyxl.styles.borders import Border, Side

def re_converted(file):
    '''The function save the .xls file to .xlsx'''
    print( str(file) + ".xls ФАЙЛ ПРИНЯТ ДЛЯ КОНВЕРТАЦИИ!")
    file_xls = file + ".xls"
    file_xlsx = file + ".xlsx"
    pyexcel.save_book_as(file_name=file_xls, dest_file_name=file_xlsx)
    print(file_xls + " СКОНВЕРТИРОВАН В .xlsx ")
    return file_xlsx

def format_check(file_name):
    ''' Format check'''
    #передаем файл для конвертации
    file_xlsx = re_converted(file_name)

    active_excel = load_workbook(filename=file_xlsx, data_only=True)  # getting data
    active_sheet = active_excel.active #active sheet

    #check
    if active_sheet["B4"].value == "Номенклатура":
        if active_sheet["D4"].value == "ХарактеристикаНоменклатуры":
            if active_sheet["G4"].value == "Единица":
                if active_sheet["J4"].value == "НоменклатураТипИзделия":
                    if active_sheet["K4"].value == "Штрихкод":
                        if active_sheet["L4"].value == "ЦенаРозничная":
                            if active_sheet["M4"].value == "Количество":
                                return True

def rezult_check():
    '''Print rezult'''
    name = input("ВВЕДИТЕ НАЗВАНИЕ ФАЙЛА: ")
    name_xlsx = name + ".xlsx"
    r = format_check(name)

    if r == True:
        print("ФАЙЛ ПРИНЯТ И СООТВЕТСТВУЕТ ФОРМАТУ!")

        active_excel = load_workbook(filename=name_xlsx, data_only=True)  # getting data
        active_sheet = active_excel.active  # active sheet

        #max row
        max_row = active_sheet.max_row

        #take rows and append date
        tip_izdeliya = [] # J
        nomenklatura = [] # B
        harakteristika = [] # D
        shtrihkod = [] # K
        edinitca = [] # G
        kolichestvo = [] # M
        chena_postypleniya = [] # O
        nds = [20] #const
        chena_roznichnaya = [] # L
        spisok_all = [tip_izdeliya,tip_izdeliya,nomenklatura,harakteristika,shtrihkod,edinitca,kolichestvo,chena_postypleniya,nds,chena_roznichnaya]

        for i in range(5,(max_row - 1)):
            j = active_sheet["J" + str(i)].value
            tip_izdeliya.append(j)
            b = active_sheet["B" + str(i)].value
            nomenklatura.append(b)
            d = active_sheet["D" + str(i)].value
            harakteristika.append(d)
            k = active_sheet["K" + str(i)].value
            shtrihkod.append(int(k))
            g = active_sheet["G" + str(i)].value
            edinitca.append(g)
            m = active_sheet["M" + str(i)].value
            kolichestvo.append(m)
            o = active_sheet["O" + str(i)].value
            chena_postypleniya.append(o)
            l = active_sheet["L" + str(i)].value
            chena_roznichnaya.append(l)
            const = 20
            nds.append(const)

        #write lists
        filename = "ШаблонЗагрузки.xlsx"
        active_excel_1 = load_workbook(filename=filename, data_only=True)  # getting data
        active_sheet_1 = active_excel_1.active  # active sheet

        #write first part in document
        #style
        style_1 = Font(name='Arial', color=colors.BLACK,
                     bold=False, size=14)  # underline='double'
        style_2 = Font(name='TimesNewRoman', color=colors.BLACK,
                     bold=False, size=14)  # underline='double'

        #A
        active_sheet_1["A1"] = 'Группа номенклатуры'
        active_sheet_1["A2"] = 'Поставщик'
        active_sheet_1["A3"] = '1'
        #B
        active_sheet_1["B1"] = 'Вид номенклатуры'
        active_sheet_1["B2"] = 'Наименование'
        active_sheet_1["B3"] = '2'
        #C
        active_sheet_1["C1"] = ' '
        active_sheet_1["C2"] = 'Наименование'
        active_sheet_1["C3"] = '3'
        #D
        active_sheet_1["D1"] = 'Характеристика'
        active_sheet_1["D2"] = 'Размер'
        active_sheet_1["D3"] = '4'
        #E
        active_sheet_1["E1"] = ' '
        active_sheet_1["E2"] = 'Штрихкод'
        active_sheet_1["E3"] = '5'
        #F
        active_sheet_1["F1"] = ' '
        active_sheet_1["F2"] = 'Ед. Измерения'
        active_sheet_1["F3"] = '6'
        #G
        active_sheet_1["G1"] = ' '
        active_sheet_1["G2"] = 'Количество'
        active_sheet_1["G3"] = '7'
        #H
        active_sheet_1["H1"] = 'Закуп'
        active_sheet_1["H2"] = 'Цена'
        active_sheet_1["H3"] = '8'
        #I
        active_sheet_1["I1"] = ' '
        active_sheet_1["I2"] = 'НДС (20%)'
        active_sheet_1["I3"] = '9'
        #J
        active_sheet_1["J1"] = ' '
        active_sheet_1["J2"] = 'Цена розничная'
        active_sheet_1["J3"] = '10'

        #style for first part in document
        for r in range (1,4):
            a = active_sheet_1["A" + str(r)]
            b = active_sheet_1["B" + str(r)]
            c = active_sheet_1["C" + str(r)]
            d = active_sheet_1["D" + str(r)]
            e = active_sheet_1["E" + str(r)]
            f = active_sheet_1["F" + str(r)]
            g = active_sheet_1["G" + str(r)]
            h = active_sheet_1["H" + str(r)]
            i = active_sheet_1["I" + str(r)]
            j = active_sheet_1["J" + str(r)]
            #style
            a.font = style_1
            b.font = style_1
            c.font = style_1
            d.font = style_1
            e.font = style_1
            f.font = style_1
            g.font = style_1
            h.font = style_1
            i.font = style_1
            j.font = style_1
            #center
            a.alignment = Alignment(horizontal='center')
            b.alignment = Alignment(horizontal='center')
            c.alignment = Alignment(horizontal='center')
            d.alignment = Alignment(horizontal='center')
            e.alignment = Alignment(horizontal='center')
            f.alignment = Alignment(horizontal='center')
            g.alignment = Alignment(horizontal='center')
            h.alignment = Alignment(horizontal='center')
            i.alignment = Alignment(horizontal='center')
            j.alignment = Alignment(horizontal='center')

        #width column
        column = 1
        while column < 11:
            i = get_column_letter(column)
            active_sheet_1.column_dimensions[i].width = 30
            column += 1

        #max row in active_excel_1
        max_row_1 = active_sheet_1.max_row

        for g in range(len(tip_izdeliya)):
            for h in range(len(spisok_all)):
                _=active_sheet_1.cell(column=h+1, row=max_row_1+g+1, value=spisok_all[h][g])

        #max row before write for format
        max_row_2 = active_sheet_1.max_row

        #number format
        for s in range(4, max_row_2 + 1):
            active_sheet_1["E" + str(s)].number_format = '#,##0'
            active_sheet_1["H" + str(s)].number_format='#,##0.00'
            active_sheet_1["J" + str(s)].number_format='#,##0.00'

        # style for second part in document
        for r in range (4,max_row_2 + 1):
            a = active_sheet_1["A" + str(r)]
            b = active_sheet_1["B" + str(r)]
            c = active_sheet_1["C" + str(r)]
            d = active_sheet_1["D" + str(r)]
            e = active_sheet_1["E" + str(r)]
            f = active_sheet_1["F" + str(r)]
            g = active_sheet_1["G" + str(r)]
            h = active_sheet_1["H" + str(r)]
            i = active_sheet_1["I" + str(r)]
            j = active_sheet_1["J" + str(r)]
            #style
            a.font = style_2
            b.font = style_2
            c.font = style_2
            d.font = style_2
            e.font = style_2
            f.font = style_2
            g.font = style_2
            h.font = style_2
            i.font = style_2
            j.font = style_2

        #border in all document
        thick_border=Border(left=Side(style='thin'),
                            right=Side(style='thin'),
                            top=Side(style='thin'),
                            bottom=Side(style='thin'))
        for r in range (1,max_row_2 + 1):
            active_sheet_1["A" + str(r)].border = thick_border
            active_sheet_1["B" + str(r)].border = thick_border
            active_sheet_1["C" + str(r)].border = thick_border
            active_sheet_1["D" + str(r)].border = thick_border
            active_sheet_1["E" + str(r)].border = thick_border
            active_sheet_1["F" + str(r)].border = thick_border
            active_sheet_1["G" + str(r)].border = thick_border
            active_sheet_1["H" + str(r)].border = thick_border
            active_sheet_1["I" + str(r)].border = thick_border
            active_sheet_1["J" + str(r)].border = thick_border

        try:
            active_excel_1.save("ШаблонЗагрузки.xlsx")
        except PermissionError:
            print("ЗАКРОЙТЕ ШАБЛОН ЗАГРУЗКИ!")
            quit()

        print("ДАННЫЕ ЗАПИСАНЫ!")

    else:
        print("ФАЙЛ НЕ СООТВЕТСТВУЕТ ФОРМАТУ!")

while 1 == 1:
    w = input("ЖЕЛАЕТЕ ЗАПИСАТЬ ФАЙЛ? (да/нет) ")
    if w == "да":
        rezult_check()
    else:
        quit()

input()

#pyinstaller -F IR_Excel_3.py